# This file will process the standardised credential data and convert it into the appropriate WHED Codes

import pymysql, cryptography
from dotenv import load_dotenv
import os, tempfile
from openpyxl import load_workbook, Workbook

def main(masterlist_path, output_path):
    # Create a new dict of institutions
    insts = get_insts(output_path)

    # Get list of FOS Codes and FOS Levels / Display Categories from WHED (or spreadsheet)
    # Open connection to the WHED
    conn = whed_connect()
    cursor = conn.cursor()
    # test connection
    cursor.execute("SELECT GlobalID, OrgName FROM whed_org LIMIT 20;")

    # Open masterlist
    if not os.path.exists(masterlist_path):
        print(f"Masterlist not found at {masterlist_path}")
        return
    
    #Read list
    print("Opening masterlist be patient...", flush = True)
    wb = load_workbook(masterlist_path)

    # Open whed_levels sheet
    whed_levels = wb['whed_levels']
  

    # get list of whed credentials from whed_levels
    whed_creds = get_creds(whed_levels, cursor)


    # get list of fos in WHED from workbook
    #TODO get the fos into the workbook and synced and update the template
    whed_fos = "spaghetti"

    # open courses sheet
    ext_cred = wb['ext_cred']

    dual_qual_count = 0

    for inst in insts:
        # skip institition if it hasn't been categorised as "confirmed"
        if inst["status"] != "confirmed":
            continue
        # For each credential in sheet
        for row in ext_cred.iter_rows(min_row=2, values_only = True):


            # If credential's institution matches the one in the loop & is not expired
            expired = str(row[2])
            if expired == "No" and inst["ext_id"] == str(row[0]):

                # Assign row data to variables
                temp_degree = {
                    "course_level": str(row[3]),
                    "course_name": str(row[4]),
                    "fos_level_1": str(row[5]),
                    "fos_level_2": str(row[6]),
                    "fos_level_3": str(row[7]),
                    "dual_qualification": str(row[23])
                }
                # increment variable if there is a dual qualification
                if temp_degree["dual_qualification"] == "Yes":
                    dual_qual_count += 1
                
                #TODO Match whed credential code
                cred_code = get_cred_code(whed_creds, temp_degree)


                #TODO Match Field to appropriate whed FOS using the following hierarchy
                    # If any of the FOS fields match, use that
                        # Get WHED FOS Code and add it to a dict
                    # If a shaved version of the credential name matches a WHED FOS field
                        # Get WHED FOS Code and add it to a dict
                    # If there is a fuzzy match
                        # Add the cred to the "to be sorted" category, and add to a bucket
                        # By bucket I mean basically to have all unsorted categories matched together, so there could potentially be 100 instances of a
                        # non-matched field (e.g. Mobile Programming) that could then be categorised by a Data Officer at the end of the program
                fos_code = get_fos_code(whed_fos, temp_degree)


                # Assign the degree (cred (bachelor, masters) + field of study (compsci, history)) to a variable
                degree = {
                    "org_id": inst["whed_id"],
                    "cred_code": cred_code,
                    "fos_code": fos_code,
                    "course_name": temp_degree["course_name"]
                }


    print(f"There were {dual_qual_count} valid dual qualifications processed")
    exit

def whed_test_connect(query):
    load_dotenv()
    timeout = 10
    connection = pymysql.connect(
        charset="utf8mb4",
        connect_timeout=timeout,
        cursorclass=pymysql.cursors.DictCursor,
        db=os.getenv("DB_NAME"),
        host=os.getenv("DB_HOST"),
        password=os.getenv("DB_PASSWORD"),
        read_timeout=timeout,
        port=int(os.getenv("DB_PORT", 3306)),
        user=os.getenv("DB_USER"),
        write_timeout=timeout,
    )
    try:
        cursor = connection.cursor()
        cursor.execute(query)
        print(cursor.fetchall())
    finally:
        connection.close()

# will return the conn for the database connection
def whed_connect():
    load_dotenv()
    timeout = 10
    connection = pymysql.connect(
        charset="utf8mb4",
        connect_timeout=timeout,
        cursorclass=pymysql.cursors.DictCursor,
        db=os.getenv("DB_NAME"),
        host=os.getenv("DB_HOST"),
        password=os.getenv("DB_PASSWORD"),
        read_timeout=timeout,
        port=int(os.getenv("DB_PORT", 3306)),
        user=os.getenv("DB_USER"),
        write_timeout=timeout,
    )
    return connection

# takes the current row of the credentials table and the whed_levels sheet to try to return the credential code
def get_cred_code(row, whed_creds):
   


    return "1138"

def get_fos_code(row, temp_degree):
    print("FOS Code Row: ", row)
    return "1111"

def get_insts(output_path):
    # initialise insts list
    insts = []
    # open output file
    if not os.path.exists(output_path):
        print(f"Output not found, did you run the categorisation on the masterlist? path: {output_path}")
        exit()

    # Read output
    print("Opening output", flush = True)
    wb = load_workbook(output_path)
    
    # Open output sheet
    ws = wb['Sheet']

    # Loop through rows
    for row in ws.iter_rows(min_row=2, values_only = True):
        # create new inst using row info
        inst = {
            "whed_id": row[3],
            "whed_name": str(row[8]),
            "whed_name_eng": str(row[6]),
            "ext_id": str(row[1]),
            "ext_name": str(row[2]),
            "ext_trading": str(row[0]),
            "status": str(row[5]),
            "match_type": str(row[7])
            }
        insts.append(inst)

    return insts

def get_creds(whed_levels, cursor):
    whed_creds = []
    for row in whed_levels.iter_rows(min_row=2, values_only = True):
        cred_name = str(row[0])
        cred_level_code = str(row[1])
        country_code = str(row[2])

        # get country ID
        country_id = cursor.execute(f"SELECT StateID FROM whed_state WHERE CountryCode = '{country_code}'")

        print(f"Country ID for Country Code ({country_code}): {country_id}")

        whed_cred = {
            "cred_name": cred_name,
            "cred_code": cred_level_code
        }
        whed_creds.append(whed_cred)
    return whed_creds
    