from openpyxl import load_workbook, Workbook #type: ignore
import os
import re


# CRICOS is an Australian Credential Registry that does not have an API, they do however regularly export from their database and release exports to the public as excel spreadsheets.
# The World Higher Education Database (WHED) aims to catalogue every higher education institution in the world including institutional and credential information.

# The purpose of this script is to process the data within the CRICOS exports and import it directly into the WHED to semi-automate the update process for Australia.


# Before the process begins the latest list of WHED-recognised institutions in Australia is exported with their ID and added as a sheet in the masterlist

# The main method is called by the main.py script
def main(masterlist_path, postgrad_codes):
    # open masterlist
    if not os.path.exists(masterlist_path):
        print(f"File not found {masterlist_path}")
        return

    # Read masterlist
    print("Opening masterlist be patient this takes a sec...", flush = True)
    wb = load_workbook(masterlist_path)
    

    # open the whed_levels sheet
    whed_levels = wb['whed_levels']

    print("Checking postgrad degrees")
    # Get the list of Postgrad Degrees
    postgrad_list = get_postgrad_list(postgrad_codes, whed_levels)

    print(f"List of postgrad credentials offered in this country:\n{postgrad_list}")  

    # Categorise external institutions
    insts = process_input(wb, postgrad_list)

    print_summary(insts)



    # write institutions to excel format to allow for later verification by Data Officers
    write_output(insts)

def get_postgrad_list(postgrad_codes, whed_levels):

    # for row of credential name

    for row in whed_levels.iter_rows(min_row=2, values_only = True):

        cred_name = str(row[0])

        level_code= str(row[1])

        if level_code in postgrad_codes:

            # append to the list of NQF codes in case the source spreadsheet uses those instead of names

            postgrad_codes.append(cred_name)

    return postgrad_codes

# Assess external institutions for WHED eligibility, categorise all institutions in external Dataset
# Takes the list of postgrad degrees and the workbook as arguments
def process_input(wb, postgrad_list):
    insts = []

    # Open the CRICOS institution sheet
    ext_inst = wb['ext_inst']
    
    # Open the courses sheet
    ext_cred = wb['ext_cred']

    # Open the WHED institution sheet
    whed_inst = wb['whed_inst']


    # For each institution
    for row in ext_inst.iter_rows(min_row=2, values_only = True):
        row_index = row[0]
        # insts are by default ineligible 
        inst = {
        "whed_id": None,
        "whed_name": None,
        "whed_name_eng": None,
        "ext_id": str(row[1]),
        "ext_name": str(row[2]).lower(),
        "ext_trading": str(row[3]).lower(),
        "status": "ineligible",
        "whed_status": None,
        "match_type": None,
        "ext_url": tidy_url(str(row[4])),
        "ext_address": str(row[5])
        }

        inst = candidate_check(inst, postgrad_list, ext_cred)

        print(f"Row {row_index}")
        print(inst['ext_name'], flush = True)

        # check whether the institution is in the WHED and update the dict as appropriate
        inst = whed_check(inst, whed_inst)

        
        print(f"Candidate status: {inst['status']}\n\n")

        insts.append(inst)

    #TODO Rename this function to something
    verify_whed_insts()

    return insts

# Checks the credentials offered at a specific institution to see if it is a possible WHED candidate, takes the institution dict, the cred list, and the cred ws as input
def candidate_check(inst, cred_list, ext_cred):
    # loop through all credentials in the list
    for row in ext_cred.iter_rows(min_row = 2, values_only = True):
        ext_id = str(row[0])
        # If the current credential is offered at the institution being checked, get the credential type
        if inst["ext_id"] == ext_id:
            cred_type = str(row[3])
            expired = str(row[2])
            # If the credential type is in the list of postgrad types and it's not expired, the institution is WHED candidate
            if cred_type in cred_list and expired == "No":
                inst["status"] = "candidate"
                return inst
    return inst


# Will try to match institutions in CRICOS to an export from the WHED and will return the instituion name, id, and match type (name, site, address) if it matches
# Takes the institution dict and the whed_institution sheet as input
def whed_check(inst, whed_inst):
    # For clarity and sanity I mapped everything to local variables
    ext_name = inst["ext_name"]
    ext_name_alt = inst["ext_trading"]
    ext_url = inst["ext_url"]
    ext_address = inst["ext_address"]

    for row in whed_inst.iter_rows(min_row=2, values_only = True):
        whed_id = str(row[0])
        
        # Have all whed names be a list that can be iterated to compare to external names
        whed_names = []
        
        whed_names.append (str(row[2]).lower())
        whed_names.append (str(row[3]).lower())
        whed_names.append (str(row[4]).lower())
        whed_url = tidy_url(str(row[5]))
        whed_address = str(row[6])


        foobar = "z"
        # check if webpages match first as it's the most definitive match
        
        # print(f"external url: {ext_url}\nwhed url: {whed_url}")
        if ext_url == whed_url:
            inst["match_type"] = "web"
            break
        # TODO: This one is not a priority yet as there should be enough for the Data Officers with name and url matches
        elif foobar == 'w':
            inst["match_type"] = "address"
            break
        # check if the external names match the whed names
        elif ext_name in whed_names or ext_name_alt in whed_names:
            inst["match_type"] = "name"
            break
        # TODO Add comparison for partial matches
    
    # If there was a match link the whed ID and check whether it should remain a WHED candidate
    if inst["match_type"] != None:
        inst["whed_id"] = whed_id
        inst["whed_name_eng"] = whed_names[1]
        if inst["status"] == "candidate":
            inst["status"] = "confirmed"
        else:
            inst["status"] = "verify"

    return inst

# Will loop through all WHED institutions and list any that were not matched, adding them to the insts dict.
# TODO Flesh out function
def verify_whed_insts():
    return 0

# Will return only the tld of a url
def tidy_url(url):
    # Remove protocol and 'www.'
    url = re.sub(r'^(https?://)?(www\.)?', '', url)
    # Get only the hostname part (before the first slash)
    url = url.split('/')[0]
    return(url)

def print_summary(insts):
    whed_candidates = 0
    whed_ineligible = 0
    whed_verify = 0
    whed_confirmed = 0

    print("----------List of ineligible institutions----------")
    for inst in insts:
        if(inst["status"] == "ineligible"):
            print(inst["ext_name"])
            whed_ineligible += 1

    print("----------List of confirmed institutions---------")
    for inst in insts:
        if(inst["status"] == "confirmed"):
            print(f"{inst['whed_name_eng']}, match type: {inst['match_type']}")
            whed_confirmed += 1

    print("----------List of Potential WHED Candidates----------")
    for inst in insts:
        if(inst["status"] == "candidate"):
            print(inst["ext_name"])
            whed_candidates += 1

    print("----------List of institutions to check validity-----")
    for inst in insts:
        if(inst["status"] == "verify"):
            print(inst["ext_name"])
            whed_verify += 1

    print("Analysis complete (for details scroll up)")
    print(f"Potential WHED level candidates: {whed_candidates}")
    print(f"Ineligible instititions based on degree offerings: {whed_ineligible}")
    print(f"Institutions in WHED that do not offer postgrad according to CRICOS: {whed_verify}")
    print(f"WHED institutions confirmed by CRICOS: {whed_confirmed}")

    print("Writing to output.xlsx, stand by")


# Write the Dicts to a spreadsheet
def write_output(insts):
    
    
    output = Workbook()
    ws = output.active
    
    # Get unique keys and add them as the column headers
    columns = list({key for row in insts for key in row.keys()})

    ws.append(columns)

    for row in insts:
        ws.append([row.get(key, "") for key in columns])
    
    
    output.save("output.xlsx")

    return 0